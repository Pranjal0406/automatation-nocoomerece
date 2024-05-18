import openpyxl

def getRowCount(file,sheetname):   # file means path
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row

def ColumnCount (file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_column

def ReadData (file,sheetname,rownum,columnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(rownum,columnum).value

def WriteData (file,sheetname,rownum,columnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(rownum,columnum).value = data
    workbook.save(file)